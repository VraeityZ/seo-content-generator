import os
import re
import pandas as pd
import anthropic
from openai import OpenAI
from datetime import datetime
import warnings
import openpyxl
import math
import logging


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

# Define output directory
OUTPUT_DIR = "output"

# Placeholder for API keys - these should be set in environment variables or Streamlit secrets
def get_api_keys(claude_api, openai_api):
    return claude_api, openai_api

# Model selection
platform = "Claude"  # @param ["Claude"]
claude_model = "claude-3-7-sonnet-latest"
chatgpt_model = "o1-mini-2024-09-12"

# Heading control variables (global for simplicity; adjust as needed for Streamlit)
h2_control = 0  # @param {"type":"number","placeholder":"0"}
h3_control = 0  # @param {"type":"number","placeholder":"0"}
h4_control = 0  # @param {"type":"number","placeholder":"0"}
h5_control = 0  # @param {"type":"number","placeholder":"0"}
h6_control = 0  # @param {"type":"number","placeholder":"0"}

# Initialize API clients
def initialize_api_clients(claude_api, openai_api):
    if platform == "Claude":
        client = anthropic.Anthropic(api_key=claude_api)
        model = claude_model
    elif platform == "ChatGPT":
        client = OpenAI(api_key=openai_api)
        model = chatgpt_model
    else:
        raise ValueError(f"Unsupported platform: {platform}")
    return client, model

##############################################################################
# UPLOAD FILE
##############################################################################
def upload_file():
    """Placeholder for Streamlit file upload."""
    return None

##############################################################################
# PARSE CORA REPORT
##############################################################################
def parse_cora_report(file_path):
    """Parses a CORA Excel report and extracts SEO requirements."""
    try:
        # Load the Excel workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Initialize default values
        primary_keyword = ""
        entities = []
        variations = []
        lsi_keywords = {}
        heading_structure = {}
        requirements = {}
        word_count = 1500  # Default
        
        # Debug info
        debug_info = {
            "sheets_found": wb.sheetnames,
            "lsi_start_row": None,
            "entities_start_row": None,
            "headings_section": None
        }
        
        # Extract requirements from CORA report
        requirements = {}
        
        # Parse "Roadmap" sheet
        if "Roadmap" in wb.sheetnames:
            roadmap_sheet = wb["Roadmap"]
            
            # Variations from A2
            raw_variations = roadmap_sheet["A2"].value
            variations = [v.strip(' "\'') for v in raw_variations.split(",") if v.strip()] if raw_variations else []
            
            # Extract requirements from "Phase 1: Title & Headings"
            marker_start = "Phase 1: Title & Headings"
            possible_end_markers = [
                "Phase 2: Content",
                "Phase 3: Authority",
                "Phase 4: Diversity",
                "Phase 6: Search Result Presentation",
                "Phase 7: Outbound Linking From the Page"
            ]
            
            # Find start row
            start_row = None
            for row in range(1, 100):
                cell_a = roadmap_sheet.cell(row=row, column=1).value
                if cell_a and marker_start in str(cell_a).strip():
                    start_row = row + 1
                    break
                    
            if start_row:
                # Find end row based on possible markers
                end_row = None
                for row in range(start_row, 100):
                    cell_a = roadmap_sheet.cell(row=row, column=1).value
                    if cell_a:
                        cell_text = str(cell_a).strip()
                        if any(marker in cell_text for marker in possible_end_markers):
                            end_row = row
                            break
                            
                if not end_row:
                    end_row = roadmap_sheet.max_row
                    
                # Extract requirements
                for row in range(start_row, end_row):
                    req_desc = roadmap_sheet.cell(row=row, column=1).value
                    req_amount_text = roadmap_sheet.cell(row=row, column=2).value
                    
                    if req_desc and req_amount_text:
                        try:
                            # Use regex to find the first number in the text
                            match = re.search(r"(\d+)", str(req_amount_text))
                            if match:
                                amount = int(match.group(1))
                                requirements[req_desc] = amount
                        except (ValueError, TypeError):
                            logging.warning(f"Could not parse requirement amount: {req_amount_text}")
                            continue
        
        # Initialize heading variables before Basic Tunings processing
        heading_2 = 0
        heading_3 = 0
        heading_4 = 0
        heading_5 = 0
        heading_6 = 0
        total_heading = 0
        
        # Initialize title and description length variables
        title_length = 60  # Default value for CP480
        desc_length = 160  # Default value for CP380
        
        # Parse "Basic Tunings" sheet
        if "Basic Tunings" in wb.sheetnames:
            basic_tunings_sheet = wb["Basic Tunings"]
            # Primary keyword from B1
            primary_keyword = basic_tunings_sheet["B1"].value.strip() if basic_tunings_sheet["B1"].value else ""
            # Word count from CP492
            for row in range(1, basic_tunings_sheet.max_row + 1):
                if basic_tunings_sheet.cell(row=row, column=2).value == "CP492":
                    word_count_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if word_count_value:
                        try:
                            word_count = int(word_count_value)
                        except ValueError:
                            pass
                    break
            # Number of H2 Tags
            for row in range(1, basic_tunings_sheet.max_row + 1):
                if basic_tunings_sheet.cell(row=row, column=2).value == "CPXR005":
                    heading_2_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if heading_2_value:
                        try:
                            heading_2 = int(heading_2_value)
                        except ValueError:
                            pass
                    break
            # Number of H3 Tags
            for row in range(1, basic_tunings_sheet.max_row + 1):
                if basic_tunings_sheet.cell(row=row, column=2).value == "CPXR006":
                    heading_3_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if heading_3_value:
                        try:
                            heading_3 = int(heading_3_value)
                        except ValueError:
                            pass
                    break
            # Number of H4 Tags
            for row in range(1, basic_tunings_sheet.max_row + 1):
                if basic_tunings_sheet.cell(row=row, column=2).value == "CPXR007":
                    heading_4_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if heading_4_value:
                        try:
                            heading_4 = int(heading_4_value)
                        except ValueError:
                            pass
                    break
            # Number of H5 Tags
            for row in range(1, basic_tunings_sheet.max_row + 1):
                if basic_tunings_sheet.cell(row=row, column=2).value == "CPXR008":
                    heading_5_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if heading_5_value:
                        try:
                            heading_5 = int(heading_5_value)
                        except ValueError:
                            pass
                    break
            # Number of H6 Tags
            for row in range(1, basic_tunings_sheet.max_row + 1):
                if basic_tunings_sheet.cell(row=row, column=2).value == "CPXR009":
                    heading_6_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if heading_6_value:
                        try:
                            heading_6 = int(heading_6_value)
                        except ValueError:
                            pass
                    break   
            requirements["Number of H2 tags"] = heading_2
            requirements["Number of H3 tags"] = heading_3
            requirements["Number of H4 tags"] = heading_4
            requirements["Number of H5 tags"] = heading_5
            requirements["Number of H6 tags"] = heading_6
            
            # Number of heading tags
            for row in range(1, basic_tunings_sheet.max_row + 1):
                if basic_tunings_sheet.cell(row=row, column=2).value == "CPXR003":
                    total_heading_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if total_heading_value:
                        try:
                            total_heading = int(total_heading_value)
                        except ValueError:
                            pass
                    break   
            
            requirements["Number of heading tags"] = total_heading

            # Extract CP480 (ideal title length) and CP380 (ideal meta description length)
            for row in range(1, basic_tunings_sheet.max_row + 1):
                cell_value = basic_tunings_sheet.cell(row=row, column=2).value
                if cell_value == "CP480":  # Meta title length
                    title_length_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if title_length_value:
                        try:
                            title_length = int(title_length_value)
                        except ValueError:
                            pass
                elif cell_value == "CP380":  # Meta description length
                    desc_length_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if desc_length_value:
                        try:
                            desc_length = int(desc_length_value)
                        except ValueError:
                            pass
            
            # Store CP480 and CP380 values in requirements
            requirements["CP480"] = title_length
            requirements["CP380"] = desc_length

        # Parse "LSI Keywords" sheet
        lsi_sheet_name = next((s for s in wb.sheetnames if "LSI" in s and "Keywords" in s), None)
        if lsi_sheet_name:
            lsi_sheet = wb[lsi_sheet_name]
            lsi_keywords_data = []
            for row in range(7, lsi_sheet.max_row + 1):  # Header at row 6
                keyword = lsi_sheet.cell(row=row, column=1).value
                avg = lsi_sheet.cell(row=row, column=2).value
                g_value = lsi_sheet.cell(row=row, column=7).value  # Column G value
                
                if keyword and avg:
                    try:
                        # Convert values to float
                        avg_float = float(avg)
                        g_float = float(g_value) if g_value else 0
                        
                        # Add to data list with G value for sorting and display
                        rounded_g = math.ceil(g_float) if g_float > 0 else 1
                        lsi_keywords_data.append((keyword, rounded_g, g_float))
                    except ValueError:
                        continue
            
            # Sort by Column G (greatest to least)
            lsi_keywords_data.sort(key=lambda x: x[2], reverse=True)
            
            # Convert to dictionary with keyword and G value for frequency (include all keywords)
            lsi_keywords = {item[0]: item[1] for item in lsi_keywords_data}
        
        # Parse "Entities" sheet
        if "Entities" in wb.sheetnames:
            entities_sheet = wb["Entities"]
            for row in range(4, entities_sheet.max_row + 1):  # Header at row 3
                entity = entities_sheet.cell(row=row, column=1).value
                if entity:
                    entities.append(str(entity).strip())
        
        # Handle heading overrides
        heading_controls = {
            2: heading_2,
            3: heading_3,
            4: heading_4,
            5: heading_5,
            6: heading_6
        }
        heading_overrides = []
        for level, control in heading_controls.items():
            if control > 0:
                for key in list(requirements.keys()):
                    if f"number of h{level} tags" in key.lower():
                        del requirements[key]
                requirements[f"Number of H{level} tags"] = control
                heading_overrides.append(f"Important: Headings override. Ignore Number of H{level} required. Instead use {control}")
        
        if "Number of Heading Tags" in requirements:
            total_headings = 1  # For H1
            for row in range(1, basic_tunings_sheet.max_row + 1):
                if basic_tunings_sheet.cell(row=row, column=2).value == "CPXR003":
                    total_heading_value = basic_tunings_sheet.cell(row=row, column=5).value
                    if total_heading_value:
                        try:
                            total_heading = int(total_heading_value)
                        except ValueError:
                            pass
                    break   
            requirements["Number of Heading Tags"] = total_heading

                # Compile results
        results = {
            "primary_keyword": primary_keyword,
            "variations": variations,
            "lsi_keywords": lsi_keywords,
            "entities": entities,
            "heading_structure": heading_structure,  # Kept for compatibility, though unused
            "requirements": requirements,
            "word_count": word_count,
            "heading_overrides": heading_overrides,
            "debug_info": debug_info
        }
        
        print(f"✅ Successfully extracted requirements for {primary_keyword}")
        return results
        
    except Exception as e:
        print(f"❌ Error parsing CORA report: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            "primary_keyword": "Sample Keyword",
            "variations": [],
            "lsi_keywords": {},
            "entities": [],
            "heading_structure": {"h2": 3, "h3": 6},
            "requirements": {},
            "word_count": 1500,
            "heading_overrides": [],
            "debug_info": {"error": str(e)}
        }

def call_claude_api(system_prompt, user_prompt, api_key, is_content_generation=False):
    """Call the Claude API with the given prompts."""
    client = anthropic.Anthropic(api_key=api_key)
    
    # Use different token budgets based on the type of generation
    max_tokens = 15000 if is_content_generation else 4000
    thinking_budget = 14000 if is_content_generation else 3500
    
    # Print debug information
    print(f"Calling Claude API:")
    print(f"Mode: {'Content Generation' if is_content_generation else 'Heading Generation'}")
    print(f"Max Tokens: {max_tokens}")
    print(f"Thinking Budget: {thinking_budget}")
    print(f"API Key: {api_key[:5]}...")
    
    # Verify prompt
    if len(user_prompt) < 50:
        print("WARNING: User prompt seems too short, might not be valid")
    
    response = client.messages.create(
        model="claude-3-7-sonnet-latest",
        max_tokens=max_tokens,
        system=system_prompt,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": user_prompt
                    }
                ]
            }
        ],
        thinking={
            "type": "enabled",
            "budget_tokens": thinking_budget
        }
    )
    # Extract content text correctly based on response structure
    # Look for the actual content, not thinking blocks
    content_text = ""
    for content_block in response.content:
        if hasattr(content_block, 'text'):
            content_text = content_block.text
            break
        elif isinstance(content_block, dict) and 'text' in content_block:
            content_text = content_block['text']
            break
    
    # Return both the response text and token usage information
    usage = {
        "input_tokens": response.usage.input_tokens,
        "output_tokens": response.usage.output_tokens,
        "total_tokens": response.usage.input_tokens + response.usage.output_tokens
    }
    return content_text, usage

##############################################################################
# GENERATE META AND HEADINGS
##############################################################################
def generate_meta_and_headings(requirements, settings=None):
    """Generate meta title, description, and heading structure based on requirements."""
    if settings is None:
        settings = {}
    
    model = settings.get('model', 'claude')
    anthropic_api_key = settings.get('anthropic_api_key', '')
    openai_api_key = settings.get('openai_api_key', '')
    
    if model == 'claude' and not anthropic_api_key:
        raise ValueError("Claude API key must be provided to use Claude")
    
    primary_keyword = requirements.get('primary_keyword', '')
    variations = requirements.get('variations', [])
    lsi_dict = requirements.get('lsi_keywords', {})
    entities = requirements.get('entities', [])
    word_count = requirements['word_count']
    
    # Get heading requirements from the requirements dictionary
    heading_structure = {
        "h2": requirements.get("requirements", {}).get("Number of H2 tags", 0),
        "h3": requirements.get("requirements", {}).get("Number of H3 tags", 0),
        "h4": requirements.get("requirements", {}).get("Number of H4 tags", 0),
        "h5": requirements.get("requirements", {}).get("Number of H5 tags", 0),
        "h6": requirements.get("requirements", {}).get("Number of H6 tags", 0)
    }
    
    # Get meta title and description length requirements if available
    # Use standard SEO recommended lengths as defaults
    meta_title_length = 60  # Recommended length for meta titles (in characters)
    meta_desc_length = 160  # Recommended length for meta descriptions (in characters)
    
    # Check if specific length requirements are provided - try multiple potential keys
    title_keys = ["title_length", "Title Length", "META TITLE LENGTH", "CP480"]
    desc_keys = ["description_length", "Description Length", "META DESCRIPTION LENGTH", "CP380"]
    
    # Try to find title length in requirements
    for key in title_keys:
        title_length_req = requirements.get("requirements", {}).get(key)
        if title_length_req:
            meta_title_length = title_length_req
            break
    
    # Try to find description length in requirements
    for key in desc_keys:
        desc_length_req = requirements.get("requirements", {}).get(key)
        if desc_length_req:
            meta_desc_length = desc_length_req
            break
    
    # Get LSI limit from requirements (default 100)
    lsi_limit = requirements.get('lsi_limit', 100)
    
    # Limit LSI keywords to top N for prompts
    if isinstance(lsi_dict, dict) and lsi_dict:
        top_lsi_keywords = sorted(lsi_dict.items(), key=lambda x: x[1], reverse=True)[:min(10, len(lsi_dict))]
        top_limit_lsi_keywords = sorted(lsi_dict.items(), key=lambda x: x[1], reverse=True)[:min(lsi_limit, len(lsi_dict))]
        lsi_formatted = "\n".join([f"- '{kw}' => at least {freq} occurrences" for kw, freq in top_lsi_keywords])
        lsi_formatted_full = "\n".join([f"- '{kw}' => at least {freq} occurrences" for kw, freq in top_limit_lsi_keywords])
    elif isinstance(lsi_dict, list) and lsi_dict:
        # Convert list to dict with frequency 1
        lsi_dict_converted = {kw: 1 for kw in lsi_dict}
        top_lsi_keywords = list(lsi_dict_converted.items())[:min(10, len(lsi_dict))]
        top_limit_lsi_keywords = list(lsi_dict_converted.items())[:min(lsi_limit, len(lsi_dict))]
        lsi_formatted = "\n".join([f"- '{kw}' => at least {freq} occurrences" for kw, freq in top_lsi_keywords])
        lsi_formatted_full = "\n".join([f"- '{kw}' => at least {freq} occurrences" for kw, freq in top_limit_lsi_keywords])
    else:
        lsi_formatted = "- No LSI keywords available\n"
        lsi_formatted_full = "- No LSI keywords available\n"
    
    # Prepare the system and user prompts
    system_prompt = """
You are a professional SEO content strategist and copywriter. Your job is to create optimized content strategies that rank well in search engines. Right now your task is to first generate a user friendly page structure utilizing the headings as specified and required by the user.
    """
    
    user_prompt_heading = f"""
Please create a meta title, meta description, and heading structure for a piece of content about "{primary_keyword}".

<requirements>
- Primary Keyword: {primary_keyword}
- Variations to consider: {', '.join(variations[:5])}
- LSI Keywords to Include:{lsi_formatted}
- Entities to Include: {', '.join(entities[:10])}
</requirements>

<step 1>
Using the information and requirements provided tackle the SEO-optimized content. First, establish the key elements required:
- Title Tag:
- Meta Description:
- Headings Tags:
Please follow these guidelines for content structure:
1. Title: Include at least one instance of the main keyword and should be within {meta_title_length} characters.
2. Meta Description: {meta_desc_length} characters or close to that length.
3. Avoid Redundancy
3A. Definition: Prevent the repetition of identical factual information, phrasing, or ideas across different sections unless necessary for context or emphasis.
3B. Guidelines:
3B1. Each section should introduce new information or a fresh perspective.
3B2. Avoid reusing the same sentences or key points under different headings.
3B3. If overlap occurs, merge sections or reframe the content to add distinct value.
3C. Example:
3C1. Redundant: Two sections both state, '[Topic] is beneficial.'
3C2. Fixed: One section defines '[Topic]', while another explains another aspect of '[Topic]'.
4. Include an FAQ if the topic involves common user questions or multiple subtopics. FAQ Section should be an H2. The Questions must each be an H3.
5. Merge variations into single headings when possible (as long as it makes sense for readability, SEO and adheres with the heading requirements).
6. IMPORTANT: Ensure and Confirm each step in the Step 1 list is met.
</step 1>

<step 2>
1. Create a heading structure with the following requirements. No Less. It can be More ONLY if absolutely necessary, otherwise no more than:
   - H1: Contains the primary keyword
   - H2: {heading_structure.get("h2", 0)} headings
   - H3: {heading_structure.get("h3", 0)} headings
   - H4: {heading_structure.get("h4", 0)} headings
   - H5: {heading_structure.get("h5", 0)} headings
   - H6: {heading_structure.get("h6", 0)} headings

2. The headings should:
   - Contain the primary keyword and/or variations where appropriate
   - Include some LSI keywords where relevant
   - Form a logical content flow
   - Be engaging and click-worthy while still being informative
   - Be formatted in Markdown (# for H1, ## for H2, etc.)
2. Confirm all the requirements are being met in the headings.
3. Confirm all the requirements are being met in the title.
4. Confirm all the requirements are being met in the description.
5.IMPORTANT: Ensure and Confirm each step in the Step 2 list is met.
</step 2>

Format your response exactly like this:
META TITLE: [Your meta title here]
META DESCRIPTION: [Your meta description here]
HEADING STRUCTURE:
[Complete markdown user journey friendly heading structure with # for H1, ## for H2, etc. Provided in order of the exact page layout eg.
# Heading 1
## Heading 2
### Heading 3
### Heading 3
## Heading 2
etc.]"""
    
    # Save the prompt to a file for reference
    with open("heading_prompt.txt", "w") as f:
        f.write(f"System Prompt:\n{system_prompt}\n\n\nUser Prompt:{user_prompt_heading}")
    
    # Print debug information
    print(f"Meta Title Length Used: {meta_title_length}")
    print(f"Meta Description Length Used: {meta_desc_length}")
    print(f"Heading Structure: {heading_structure}")
    
    # Make the API call
    if model == 'claude':
        result, token_usage = call_claude_api(system_prompt, user_prompt_heading, anthropic_api_key, is_content_generation=False)
    else:
        raise ValueError(f"Unsupported model: {model}")
    
    # Parse the result to extract meta title, description, and headings
    meta_title = ""
    meta_description = ""
    heading_structure = ""
    
    if "META TITLE:" in result:
        meta_title = result.split("META TITLE:")[1].split("META DESCRIPTION:")[0].strip()
    
    if "META DESCRIPTION:" in result:
        meta_description = result.split("META DESCRIPTION:")[1].split("HEADING STRUCTURE:")[0].strip()
    
    if "HEADING STRUCTURE:" in result:
        heading_structure = result.split("HEADING STRUCTURE:")[1].strip()
    
    return {
        "meta_title": meta_title,
        "meta_description": meta_description,
        "heading_structure": heading_structure,
        "token_usage": token_usage
    }

def generate_content_from_headings(requirements, heading_structure, settings=None):
    """Generate content based on the provided heading structure."""
    if settings is None:
        settings = {}
    
    # Print debug info about inputs
    print(f"Content Generation Starting:")
    print(f"API Key Available: {'Yes' if settings.get('anthropic_api_key') else 'No'}")
    print(f"Word Count Requested: {requirements.get('word_count', 'Not specified')}")
    print(f"LSI Limit: {requirements.get('lsi_limit', 'Not specified')}")
    print(f"Heading Structure Length: {len(heading_structure) if heading_structure else 0} chars")
    
    primary_keyword = requirements.get('primary_keyword', '')
    variations = requirements.get('variations', [])
    lsi_dict = requirements.get('lsi_keywords', {})
    entities = requirements.get('entities', [])
    word_count = requirements.get('word_count', 1500)
    
    # Format keyword variations
    variations_text = ", ".join(variations[:10]) if variations else "None"
    
    # Get LSI limit from requirements (default 100)
    lsi_limit = requirements.get('lsi_limit', 100)
    
    # Get top N LSI keywords based on user preference
    lsi_formatted_100 = ""
    if isinstance(lsi_dict, dict) and lsi_dict:
        # Sort by frequency and take top N based on lsi_limit
        top_lsi_keywords = sorted(lsi_dict.items(), key=lambda x: x[1], reverse=True)[:min(lsi_limit, len(lsi_dict))]
        lsi_formatted_100 = "\n".join([f"- '{kw}' => use at least {freq} times" for kw, freq in top_lsi_keywords])
    elif isinstance(lsi_dict, list) and lsi_dict:
        # For list format, assume frequency of 1 for each keyword
        lsi_keywords_subset = lsi_dict[:min(lsi_limit, len(lsi_dict))]
        lsi_formatted_100 = "\n".join([f"- '{kw}' => use at least 1 time" for kw in lsi_keywords_subset])
    
    if not lsi_formatted_100:
        lsi_formatted_100 = "- No LSI keywords available\n"
    
    # Format entities
    if entities:
        entities_text = "\n".join([f"- {entity}" for entity in entities[:20]])
    else:
        entities_text = "- No specific entities required"
    
    # Get meta information if available
    meta_title = requirements.get('meta_title', '')
    meta_description = requirements.get('meta_description', '')
    
    # If meta information is not in requirements, check if it's in meta_and_headings
    if not meta_title and 'meta_and_headings' in requirements:
        meta_title = requirements.get('meta_and_headings', {}).get('meta_title', '')
        meta_description = requirements.get('meta_and_headings', {}).get('meta_description', '')
    
    # Ensure heading structure is sanitized
    if not heading_structure or not heading_structure.strip():
        heading_structure = "# " + primary_keyword
    
    # Construct the system prompt
    system_prompt = """You are an expert SEO content writer with deep knowledge about creating high-quality, engaging, and optimized content."""
    
    # Construct the user prompt for content generation
    user_prompt = f"""
# SEO Content Writing Task

Please write a comprehensive, SEO-optimized article about **{primary_keyword}**. 
    
1. Meta Information (do not change or add to it):
- Meta Title: {meta_title}
- Meta Description: {meta_description}
    
2. Key Requirements:
- Word Count: {word_count} words (minimum). Must be no less than {word_count} but no more than {word_count + 100}
- Primary Keyword: {primary_keyword}
- Use the EXACT following heading structure (do not change or add to it):
<headings_structure>
{heading_structure}
</headings_structure>
    
3. Keyword Usage Requirements:
- Use the primary keyword ({primary_keyword}) in the first 100 words, in at least one H2 heading, and naturally throughout the content.
- Include these keyword variations naturally: {variations_text}
    
4. LSI Keywords to Include (with minimum frequencies):
{lsi_formatted_100}
    
5. Entities/Topics to Cover:
{entities_text}
    
6. Content Writing Guidelines:
- 1. Write in a clear, authoritative style suitable for an expert audience
- 2. Make the content deeply informative and comprehensive
- 3. Always write in active voice and maintain a conversational but professional tone
- 4. Include only factually accurate information
- 5. Ensure the content flows naturally between sections
- 6. Include the primary keyword in the first 100 words of the content
- 7. Format the content using markdown
- 8. DO NOT include any introductory notes, explanations, or meta-commentary about your process
- 9. DO NOT use placeholder text or suggest that the client should add information
- 10. DO NOT use the phrases "in conclusion" or "in summary" for the final section

IMPORTANT: Return ONLY the pure markdown content without any explanations, introductions, or notes about your approach.
"""
    
    # Save the prompt to a file for reference
    with open("content_prompt.txt", "w", encoding="utf-8") as f:
        f.write(user_prompt)
    
    # Call the API based on the settings
    if settings.get('model', '').lower() == 'claude' and settings.get('anthropic_api_key'):
        result, token_usage = call_claude_api(system_prompt, user_prompt, settings.get('anthropic_api_key'), is_content_generation=True)
    else:
        # Default to Claude if no valid settings are provided
        if settings.get('anthropic_api_key'):
            result, token_usage = call_claude_api(system_prompt, user_prompt, settings.get('anthropic_api_key'), is_content_generation=True)
        else:
            raise ValueError("No valid API key provided. Please provide either an Anthropic or OpenAI API key.")
    
    # Process the result to get clean markdown
    markdown_content = extract_markdown_content(result)
    
    # Ensure we have content
    if not markdown_content or len(markdown_content.strip()) < 100:
        # If extraction failed, use the full response but try to clean it
        markdown_content = result.strip()
    
    # Convert to HTML
    html_content = markdown_to_html(markdown_content)
    
    # Save to a file
    filename = f"seo_content_{primary_keyword.replace(' ', '_').lower()}.md"
    with open(filename, "w", encoding="utf-8") as f:
        f.write(markdown_content)
    
    # Return results as a dictionary with all necessary information
    return {
        'markdown': markdown_content,
        'html': html_content,
        'filename': filename,
        'token_usage': token_usage
    }

def generate_content(requirements, settings=None):
    """Legacy function that combines both steps for backward compatibility."""
    if settings is None:
        settings = {}
    
    # Generate meta and headings
    meta_and_headings = generate_meta_and_headings(requirements, settings)
    
    # Generate content using the headings
    markdown_content, html_content, save_path = generate_content_from_headings(
        requirements, 
        meta_and_headings["heading_structure"],
        settings
    )
    
    return markdown_content, html_content, save_path

##############################################################################
# SAVE MARKDOWN
##############################################################################
def save_markdown_to_file(markdown_str, keyword, iteration):
    """Saves markdown content to a file."""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = keyword.lower().replace(" ", "-").replace("/", "-")
    filename = f"{OUTPUT_DIR}/seo_content_{safe_keyword}_iteration_{iteration}_{timestamp}.md"
    
    with open(filename, "w", encoding="utf-8") as f:
        f.write(markdown_str)
    
    print(f"\nMarkdown content saved to: {filename}")
    return filename

##############################################################################
# EXTRACT MARKDOWN FROM RESPONSE
##############################################################################
def extract_markdown_from_response(response_text):
    """Extracts markdown content from an API response."""
    markdown_match = re.search(r'```(?:markdown)?(.*?)```', response_text, re.DOTALL)
    if markdown_match:
        return markdown_match.group(1).strip()
    return response_text.strip()

##############################################################################
# EXTRACT HTML FROM RESPONSE
##############################################################################
def extract_html_from_response(response_text):
    """Extracts HTML content from an API response."""
    html_match = re.search(r'```(?:html)?(.*?)```', response_text, re.DOTALL)
    if html_match:
        return html_match.group(1).strip()
    
    html_tag_match = re.search(r'<html.*?>(.*?)</html>', response_text, re.DOTALL)
    if html_tag_match:
        return f"<html>{html_tag_match.group(1)}</html>"
    
    return response_text.strip()

##############################################################################
# EXTRACT MARKDOWN CONTENT
##############################################################################
def extract_markdown_content(response_text):
    """
    Extracts clean markdown content from an API response.
    
    Args:
        response_text (str): The raw response text from the API
        
    Returns:
        str: Clean markdown content
    """
    # First try to find content between markdown code blocks
    markdown_match = re.search(r'```(?:markdown)?(.*?)```', response_text, re.DOTALL)
    if markdown_match:
        return markdown_match.group(1).strip()
    
    # If no code blocks, check for common formats Claude uses
    # Look for content structure markers (headings)
    heading_pattern = r'^#+ '
    
    # Return the full response text but clean up obvious assistant preambles/postambles
    lines = response_text.split("\n")
    content_lines = []
    skip_line = False
    content_started = False
    
    for line in lines:
        # Skip known preamble patterns
        if not content_started:
            if (line.strip() == "" or 
                line.strip().startswith("Here's") or 
                line.strip().startswith("I've created") or
                line.strip().startswith("Here is")):
                continue
            if re.match(heading_pattern, line.strip()) or line.strip().startswith('*') or line.strip().startswith('-'):
                content_started = True
            elif len(line.strip()) > 0:
                content_started = True
        
        # Skip known postamble patterns - only if they're isolated lines and not part of content
        if (line.strip() == "Let me know if you need any revisions." or
            line.strip() == "Let me know if you would like any changes." or
            line.strip() == "Is there anything else you'd like me to help with?"):
            skip_line = True
        
        if content_started and not skip_line:
            content_lines.append(line)
    
    return "\n".join(content_lines).strip()

##############################################################################
# MARKDOWN TO HTML
##############################################################################
def markdown_to_html(markdown_content):
    """
    Convert markdown to HTML.
    
    Args:
        markdown_content (str): Markdown content to convert
        
    Returns:
        str: HTML content
    """
    try:
        import markdown
        # Convert markdown to HTML
        converted_html = markdown.markdown(markdown_content)
        
        # Wrap in HTML document with styling
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Generated Content</title>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; max-width: 800px; margin: 0 auto; padding: 20px; }}
                h1 {{ color: #333; }}
                h2 {{ color: #444; border-bottom: 1px solid #eee; padding-bottom: 10px; }}
                h3 {{ color: #555; }}
                code {{ background-color: #f5f5f5; padding: 2px 4px; border-radius: 4px; }}
                pre {{ background-color: #f5f5f5; padding: 10px; border-radius: 4px; overflow-x: auto; }}
                blockquote {{ border-left: 4px solid #ddd; padding-left: 10px; color: #666; }}
                a {{ color: #0366d6; text-decoration: none; }}
                a:hover {{ text-decoration: underline; }}
            </style>
        </head>
        <body>
            {converted_html}
        </body>
        </html>
        """
    except ImportError:
        # Fallback if markdown library isn't available
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Generated Content</title>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; max-width: 800px; margin: 0 auto; padding: 20px; }}
            </style>
        </head>
        <body>
            <pre>{markdown_content}</pre>
        </body>
        </html>
        """
    return html

##############################################################################
# MAIN FUNCTION
##############################################################################
def main(claude_api, openai_api):
    """Orchestrates the markdown generation process."""
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        uploaded = upload_file()
        if not uploaded:
            print("No file uploaded. Exiting.")
            return
        
        file_path = uploaded
        print("Parsing CORA report...")
        requirements = parse_cora_report(file_path)
        print(f"✅ Successfully extracted requirements for {requirements['primary_keyword']}")
        print(f"Primary Keyword: {requirements['primary_keyword']}")
        print(f"Word Count Target: {requirements['word_count']}")
        print(f"Entities Found: {len(requirements['entities'])}")
        print(f"LSI Keywords Found: {len(requirements['lsi_keywords'])}")
        print()
        
        markdown_content = generate_content(requirements, claude_api)
        save_markdown_to_file(markdown_content, requirements["primary_keyword"], 1)
        return markdown_content
    except Exception as e:
        print(f"Error in main function: {e}")
        return None

if __name__ == "__main__":
    claude_api = os.environ.get("CLAUDE_API_KEY")
    openai_api = os.environ.get("OPENAI_API_KEY")
    main(claude_api, openai_api)